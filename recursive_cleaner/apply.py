"""Apply cleaning functions to data files."""

import csv
import importlib.util
import json
from pathlib import Path
from typing import Callable

from .parsers import MARKITDOWN_EXTENSIONS

# Text formats that should be converted to markdown (excludes spreadsheets)
TEXT_MARKITDOWN_EXTENSIONS = MARKITDOWN_EXTENSIONS - {".xlsx", ".xls", ".ods"}


def load_cleaning_module(functions_path: str):
    """
    Dynamically import a cleaning_functions.py file.

    Args:
        functions_path: Path to the cleaning functions file

    Returns:
        The imported module

    Raises:
        FileNotFoundError: If the functions file doesn't exist
        ImportError: If the module cannot be imported
    """
    path = Path(functions_path)
    if not path.exists():
        raise FileNotFoundError(f"Functions file not found: {functions_path}")

    spec = importlib.util.spec_from_file_location("cleaning_module", path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module from: {functions_path}")

    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def get_default_output_path(input_path: str, force_ext: str | None = None) -> str:
    """
    Generate default output path: input.cleaned.ext

    Args:
        input_path: Path to the input file
        force_ext: Override the output extension (e.g., ".xlsx" for .xls files)

    Returns:
        Path string for the output file
    """
    path = Path(input_path)
    suffix = path.suffix.lower()

    if force_ext:
        ext = force_ext
    elif suffix == ".xls":
        # .xls files are written as .xlsx
        ext = ".xlsx"
    elif suffix == ".txt" or suffix in TEXT_MARKITDOWN_EXTENSIONS:
        # Text formats output as markdown
        ext = ".md"
    else:
        ext = path.suffix

    return str(path.with_suffix(f".cleaned{ext}"))


def apply_to_jsonl(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Stream JSONL: read line, clean, write line.

    Args:
        input_path: Path to input JSONL file
        output_path: Path for output JSONL file
        clean_fn: Cleaning function to apply to each record
        on_progress: Optional progress callback

    Returns:
        Number of records processed
    """
    records_processed = 0

    with open(input_path, "r", encoding="utf-8") as infile, \
         open(output_path, "w", encoding="utf-8") as outfile:
        for line in infile:
            line = line.strip()
            if not line:
                continue

            record = json.loads(line)
            cleaned = clean_fn(record)
            outfile.write(json.dumps(cleaned) + "\n")

            records_processed += 1
            if on_progress:
                on_progress({"type": "apply_progress", "records_processed": records_processed})

    return records_processed


def apply_to_csv(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Stream CSV: DictReader to clean each row, DictWriter to output.

    Args:
        input_path: Path to input CSV file
        output_path: Path for output CSV file
        clean_fn: Cleaning function to apply to each record
        on_progress: Optional progress callback

    Returns:
        Number of records processed
    """
    records_processed = 0

    with open(input_path, "r", encoding="utf-8", newline="") as infile:
        reader = csv.DictReader(infile)
        fieldnames = reader.fieldnames

        if not fieldnames:
            return 0

        with open(output_path, "w", encoding="utf-8", newline="") as outfile:
            writer = csv.DictWriter(outfile, fieldnames=fieldnames)
            writer.writeheader()

            for row in reader:
                cleaned = clean_fn(row)
                writer.writerow(cleaned)

                records_processed += 1
                if on_progress:
                    on_progress({"type": "apply_progress", "records_processed": records_processed})

    return records_processed


def apply_to_json(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Batch JSON array: load all, clean each, write array.

    Args:
        input_path: Path to input JSON file
        output_path: Path for output JSON file
        clean_fn: Cleaning function to apply to each record
        on_progress: Optional progress callback

    Returns:
        Number of records processed
    """
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        # Single object - wrap, clean, unwrap
        cleaned = clean_fn(data)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(cleaned, f, indent=2)
        if on_progress:
            on_progress({"type": "apply_progress", "records_processed": 1})
        return 1

    cleaned_data = []
    for i, record in enumerate(data):
        cleaned = clean_fn(record)
        cleaned_data.append(cleaned)

        if on_progress:
            on_progress({"type": "apply_progress", "records_processed": i + 1})

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(cleaned_data, f, indent=2)

    return len(cleaned_data)


def apply_to_parquet(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Batch Parquet: load as list of dicts, clean each, write back.

    Args:
        input_path: Path to input Parquet file
        output_path: Path for output Parquet file
        clean_fn: Cleaning function to apply to each record
        on_progress: Optional progress callback

    Returns:
        Number of records processed

    Raises:
        ImportError: If pyarrow is not installed
    """
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError:
        raise ImportError(
            "pyarrow is required for parquet files. "
            "Install with: pip install recursive-cleaner[parquet]"
        )

    table = pq.read_table(input_path)
    records = table.to_pylist()

    cleaned_data = []
    for i, record in enumerate(records):
        cleaned = clean_fn(record)
        cleaned_data.append(cleaned)

        if on_progress:
            on_progress({"type": "apply_progress", "records_processed": i + 1})

    # Write back as parquet
    cleaned_table = pa.Table.from_pylist(cleaned_data)
    pq.write_table(cleaned_table, output_path)

    return len(cleaned_data)


def apply_to_excel(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Batch Excel: load as list of dicts, clean each, write back.

    Args:
        input_path: Path to input Excel file (.xlsx or .xls)
        output_path: Path for output Excel file (.xlsx)
        clean_fn: Cleaning function to apply to each record
        on_progress: Optional progress callback

    Returns:
        Number of records processed

    Raises:
        ImportError: If openpyxl (or xlrd for .xls) is not installed
    """
    suffix = Path(input_path).suffix.lower()

    if suffix == ".xls":
        # Use xlrd for .xls files
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "xlrd is required for .xls files. "
                "Install with: pip install recursive-cleaner[excel]"
            )

        workbook = xlrd.open_workbook(input_path)
        sheet = workbook.sheet_by_index(0)

        if sheet.nrows < 1:
            return 0

        # First row is headers
        headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        records = []
        for row_idx in range(1, sheet.nrows):
            row_data = {}
            for col_idx, header in enumerate(headers):
                row_data[header] = sheet.cell_value(row_idx, col_idx)
            records.append(row_data)
    else:
        # Use openpyxl for .xlsx files
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for .xlsx files. "
                "Install with: pip install recursive-cleaner[excel]"
            )

        workbook = load_workbook(input_path, read_only=True)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return 0

        # First row is headers
        headers = [str(h) if h is not None else "" for h in rows[0]]
        records = []
        for row in rows[1:]:
            row_data = {}
            for col_idx, header in enumerate(headers):
                value = row[col_idx] if col_idx < len(row) else None
                row_data[header] = value
            records.append(row_data)

        workbook.close()

    # Clean records
    cleaned_data = []
    for i, record in enumerate(records):
        cleaned = clean_fn(record)
        cleaned_data.append(cleaned)

        if on_progress:
            on_progress({"type": "apply_progress", "records_processed": i + 1})

    # Write back as xlsx using openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for writing Excel files. "
            "Install with: pip install recursive-cleaner[excel]"
        )

    wb = Workbook()
    ws = wb.active

    if cleaned_data:
        # Write headers
        fieldnames = list(cleaned_data[0].keys())
        ws.append(fieldnames)

        # Write data rows
        for record in cleaned_data:
            ws.append([record.get(k) for k in fieldnames])

    wb.save(output_path)

    return len(cleaned_data)


def apply_to_text(
    input_path: str,
    output_path: str,
    clean_fn: Callable,
    on_progress: Callable[[dict], None] | None = None,
) -> int:
    """
    Process text/document files: extract text, clean, write as markdown.

    Args:
        input_path: Path to input file (.txt or markitdown format)
        output_path: Path for output markdown file
        clean_fn: Cleaning function to apply to the text
        on_progress: Optional progress callback

    Returns:
        Number of records processed (always 1 for text)

    Raises:
        ImportError: If markitdown is not installed (for non-.txt files)
    """
    suffix = Path(input_path).suffix.lower()

    if suffix == ".txt":
        # Plain text - read directly
        with open(input_path, "r", encoding="utf-8") as f:
            content = f.read()
    else:
        # Use markitdown for other formats
        try:
            from markitdown import MarkItDown
        except ImportError:
            raise ImportError(
                "markitdown is required for this file type. "
                "Install with: pip install recursive-cleaner[markitdown]"
            )

        md = MarkItDown()
        result = md.convert(input_path)
        content = result.text_content

    # Clean the text content
    cleaned = clean_fn(content)

    # Write as markdown
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(cleaned)

    if on_progress:
        on_progress({"type": "apply_progress", "records_processed": 1})

    return 1


def apply_cleaning(
    input_path: str,
    functions_path: str,
    output_path: str | None = None,
    on_progress: Callable[[dict], None] | None = None,
) -> str:
    """
    Apply cleaning functions to a data file.

    Args:
        input_path: Path to input data file
        functions_path: Path to cleaning_functions.py
        output_path: Path for output file (default: input.cleaned.ext)
        on_progress: Optional progress callback

    Returns:
        Path to output file

    Raises:
        FileNotFoundError: If input or functions file not found
        ImportError: If functions file cannot be imported
        ValueError: If input format is unsupported
    """
    # Validate input file exists
    if not Path(input_path).exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load cleaning module
    module = load_cleaning_module(functions_path)

    # Get the clean_data function
    if not hasattr(module, "clean_data"):
        raise ImportError(f"Functions file missing clean_data() function: {functions_path}")

    clean_fn = module.clean_data

    # Determine output path
    suffix = Path(input_path).suffix.lower()
    if output_path is None:
        output_path = get_default_output_path(input_path)

    # Route by format
    format_handlers = {
        ".jsonl": apply_to_jsonl,
        ".csv": apply_to_csv,
        ".json": apply_to_json,
        ".parquet": apply_to_parquet,
        ".xlsx": apply_to_excel,
        ".xls": apply_to_excel,
    }

    handler = format_handlers.get(suffix)

    # Check for text formats (.txt and markitdown extensions, excluding spreadsheets)
    if handler is None:
        if suffix == ".txt" or suffix in TEXT_MARKITDOWN_EXTENSIONS:
            handler = apply_to_text

    if handler is None:
        raise ValueError(f"Unsupported format: {suffix}")

    # Emit start event (total_records unknown for streaming formats)
    if on_progress:
        on_progress({"type": "apply_start", "total_records": None})

    # Apply cleaning
    total_records = handler(input_path, output_path, clean_fn, on_progress)

    # Emit complete event
    if on_progress:
        on_progress({
            "type": "apply_complete",
            "total_records": total_records,
            "output_path": output_path,
        })

    return output_path
