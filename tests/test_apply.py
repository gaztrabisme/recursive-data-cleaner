"""Tests for apply module."""

import csv
import json
import tempfile
from pathlib import Path

import pytest

from recursive_cleaner.apply import (
    apply_cleaning,
    apply_to_csv,
    apply_to_json,
    apply_to_jsonl,
    get_default_output_path,
    load_cleaning_module,
)


# --- Fixtures ---


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


@pytest.fixture
def cleaning_functions_file(temp_dir):
    """Create a simple cleaning_functions.py file."""
    code = '''
def normalize_name(record):
    """Normalize the name field to uppercase."""
    if "name" in record:
        record["name"] = record["name"].upper()
    return record

def clean_data(record):
    """Main cleaning entry point."""
    record = normalize_name(record)
    return record
'''
    path = temp_dir / "cleaning_functions.py"
    path.write_text(code)
    return path


@pytest.fixture
def jsonl_file(temp_dir):
    """Create a test JSONL file."""
    path = temp_dir / "data.jsonl"
    records = [
        {"name": "alice", "age": 30},
        {"name": "bob", "age": 25},
        {"name": "charlie", "age": 35},
    ]
    with open(path, "w") as f:
        for r in records:
            f.write(json.dumps(r) + "\n")
    return path


@pytest.fixture
def csv_file(temp_dir):
    """Create a test CSV file."""
    path = temp_dir / "data.csv"
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "age"])
        writer.writeheader()
        writer.writerow({"name": "alice", "age": "30"})
        writer.writerow({"name": "bob", "age": "25"})
    return path


@pytest.fixture
def json_file(temp_dir):
    """Create a test JSON file."""
    path = temp_dir / "data.json"
    data = [
        {"name": "alice", "age": 30},
        {"name": "bob", "age": 25},
    ]
    with open(path, "w") as f:
        json.dump(data, f)
    return path


# --- Test load_cleaning_module ---


class TestLoadCleaningModule:
    def test_loads_valid_module(self, cleaning_functions_file):
        module = load_cleaning_module(str(cleaning_functions_file))
        assert hasattr(module, "clean_data")
        assert callable(module.clean_data)

    def test_raises_file_not_found(self, temp_dir):
        with pytest.raises(FileNotFoundError):
            load_cleaning_module(str(temp_dir / "nonexistent.py"))

    def test_clean_data_works(self, cleaning_functions_file):
        module = load_cleaning_module(str(cleaning_functions_file))
        result = module.clean_data({"name": "test"})
        assert result["name"] == "TEST"


# --- Test get_default_output_path ---


class TestGetDefaultOutputPath:
    def test_basic_path(self):
        result = get_default_output_path("/path/to/data.jsonl")
        assert result == "/path/to/data.cleaned.jsonl"

    def test_csv_path(self):
        result = get_default_output_path("data.csv")
        assert result == "data.cleaned.csv"

    def test_force_ext(self):
        result = get_default_output_path("data.xls", force_ext=".xlsx")
        assert result == "data.cleaned.xlsx"


# --- Test apply_to_jsonl ---


class TestApplyToJsonl:
    def test_basic_apply(self, jsonl_file, temp_dir):
        output = temp_dir / "output.jsonl"
        clean_fn = lambda r: {**r, "name": r["name"].upper()}

        count = apply_to_jsonl(str(jsonl_file), str(output), clean_fn)

        assert count == 3
        assert output.exists()

        with open(output) as f:
            lines = f.readlines()
        assert len(lines) == 3
        assert json.loads(lines[0])["name"] == "ALICE"

    def test_progress_callback(self, jsonl_file, temp_dir):
        output = temp_dir / "output.jsonl"
        events = []

        count = apply_to_jsonl(
            str(jsonl_file),
            str(output),
            lambda r: r,
            on_progress=lambda e: events.append(e),
        )

        assert count == 3
        assert len(events) == 3
        assert events[0]["records_processed"] == 1
        assert events[2]["records_processed"] == 3

    def test_skips_empty_lines(self, temp_dir):
        input_file = temp_dir / "data.jsonl"
        input_file.write_text('{"a": 1}\n\n{"b": 2}\n')
        output = temp_dir / "output.jsonl"

        count = apply_to_jsonl(str(input_file), str(output), lambda r: r)
        assert count == 2


# --- Test apply_to_csv ---


class TestApplyCsv:
    def test_basic_apply(self, csv_file, temp_dir):
        output = temp_dir / "output.csv"
        clean_fn = lambda r: {**r, "name": r["name"].upper()}

        count = apply_to_csv(str(csv_file), str(output), clean_fn)

        assert count == 2
        assert output.exists()

        with open(output, newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert len(rows) == 2
        assert rows[0]["name"] == "ALICE"

    def test_preserves_headers(self, csv_file, temp_dir):
        output = temp_dir / "output.csv"

        apply_to_csv(str(csv_file), str(output), lambda r: r)

        with open(output, newline="") as f:
            reader = csv.DictReader(f)
            assert reader.fieldnames == ["name", "age"]

    def test_progress_callback(self, csv_file, temp_dir):
        output = temp_dir / "output.csv"
        events = []

        apply_to_csv(
            str(csv_file),
            str(output),
            lambda r: r,
            on_progress=lambda e: events.append(e),
        )

        assert len(events) == 2


# --- Test apply_to_json ---


class TestApplyJson:
    def test_basic_apply_array(self, json_file, temp_dir):
        output = temp_dir / "output.json"
        clean_fn = lambda r: {**r, "name": r["name"].upper()}

        count = apply_to_json(str(json_file), str(output), clean_fn)

        assert count == 2
        with open(output) as f:
            data = json.load(f)
        assert len(data) == 2
        assert data[0]["name"] == "ALICE"

    def test_single_object(self, temp_dir):
        input_file = temp_dir / "single.json"
        input_file.write_text('{"name": "alice"}')
        output = temp_dir / "output.json"

        count = apply_to_json(str(input_file), str(output), lambda r: {**r, "name": r["name"].upper()})

        assert count == 1
        with open(output) as f:
            data = json.load(f)
        assert data["name"] == "ALICE"

    def test_progress_callback(self, json_file, temp_dir):
        output = temp_dir / "output.json"
        events = []

        apply_to_json(
            str(json_file),
            str(output),
            lambda r: r,
            on_progress=lambda e: events.append(e),
        )

        assert len(events) == 2


# --- Test apply_cleaning (main entry point) ---


class TestApplyCleaning:
    def test_jsonl_end_to_end(self, jsonl_file, cleaning_functions_file, temp_dir):
        output = temp_dir / "output.jsonl"

        result = apply_cleaning(
            str(jsonl_file),
            str(cleaning_functions_file),
            str(output),
        )

        assert result == str(output)
        assert output.exists()
        with open(output) as f:
            first_line = json.loads(f.readline())
        assert first_line["name"] == "ALICE"

    def test_csv_end_to_end(self, csv_file, cleaning_functions_file, temp_dir):
        output = temp_dir / "output.csv"

        result = apply_cleaning(
            str(csv_file),
            str(cleaning_functions_file),
            str(output),
        )

        assert result == str(output)
        with open(output, newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert rows[0]["name"] == "ALICE"

    def test_json_end_to_end(self, json_file, cleaning_functions_file, temp_dir):
        output = temp_dir / "output.json"

        result = apply_cleaning(
            str(json_file),
            str(cleaning_functions_file),
            str(output),
        )

        assert result == str(output)
        with open(output) as f:
            data = json.load(f)
        assert data[0]["name"] == "ALICE"

    def test_default_output_path(self, jsonl_file, cleaning_functions_file):
        result = apply_cleaning(
            str(jsonl_file),
            str(cleaning_functions_file),
        )

        expected = str(jsonl_file).replace(".jsonl", ".cleaned.jsonl")
        assert result == expected
        assert Path(result).exists()
        # Cleanup
        Path(result).unlink()

    def test_progress_events(self, jsonl_file, cleaning_functions_file, temp_dir):
        output = temp_dir / "output.jsonl"
        events = []

        apply_cleaning(
            str(jsonl_file),
            str(cleaning_functions_file),
            str(output),
            on_progress=lambda e: events.append(e),
        )

        types = [e["type"] for e in events]
        assert types[0] == "apply_start"
        assert types[-1] == "apply_complete"
        assert events[-1]["total_records"] == 3

    def test_file_not_found(self, cleaning_functions_file, temp_dir):
        with pytest.raises(FileNotFoundError, match="Input file not found"):
            apply_cleaning(
                str(temp_dir / "nonexistent.jsonl"),
                str(cleaning_functions_file),
            )

    def test_functions_file_not_found(self, jsonl_file, temp_dir):
        with pytest.raises(FileNotFoundError, match="Functions file not found"):
            apply_cleaning(
                str(jsonl_file),
                str(temp_dir / "nonexistent.py"),
            )

    def test_missing_clean_data_function(self, jsonl_file, temp_dir):
        bad_file = temp_dir / "bad.py"
        bad_file.write_text("def other_function(): pass")

        with pytest.raises(ImportError, match="missing clean_data"):
            apply_cleaning(str(jsonl_file), str(bad_file))

    def test_unsupported_format(self, cleaning_functions_file, temp_dir):
        weird_file = temp_dir / "data.weird"
        weird_file.write_text("data")

        with pytest.raises(ValueError, match="Unsupported format"):
            apply_cleaning(str(weird_file), str(cleaning_functions_file))


# --- Test CLI apply command ---


class TestCLIApply:
    def test_cli_apply_jsonl(self, jsonl_file, cleaning_functions_file, temp_dir):
        import subprocess

        output = temp_dir / "output.jsonl"
        result = subprocess.run(
            ["python", "-m", "recursive_cleaner", "apply",
             str(jsonl_file), "-f", str(cleaning_functions_file), "-o", str(output)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        assert "Cleaned data written to" in result.stdout
        assert output.exists()

    def test_cli_apply_file_not_found(self, cleaning_functions_file, temp_dir):
        import subprocess

        result = subprocess.run(
            ["python", "-m", "recursive_cleaner", "apply",
             str(temp_dir / "nonexistent.jsonl"), "-f", str(cleaning_functions_file)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 1
        assert "not found" in result.stderr.lower()

    def test_cli_apply_functions_not_found(self, jsonl_file, temp_dir):
        import subprocess

        result = subprocess.run(
            ["python", "-m", "recursive_cleaner", "apply",
             str(jsonl_file), "-f", str(temp_dir / "nonexistent.py")],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 1
        assert "not found" in result.stderr.lower()

    def test_cli_apply_import_error(self, jsonl_file, temp_dir):
        import subprocess

        bad_file = temp_dir / "bad.py"
        bad_file.write_text("def other(): pass")

        result = subprocess.run(
            ["python", "-m", "recursive_cleaner", "apply",
             str(jsonl_file), "-f", str(bad_file)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 2
        assert "clean_data" in result.stderr.lower()

    def test_cli_apply_default_output(self, jsonl_file, cleaning_functions_file):
        import subprocess

        result = subprocess.run(
            ["python", "-m", "recursive_cleaner", "apply",
             str(jsonl_file), "-f", str(cleaning_functions_file)],
            capture_output=True,
            text=True,
        )

        assert result.returncode == 0
        expected_output = str(jsonl_file).replace(".jsonl", ".cleaned.jsonl")
        assert Path(expected_output).exists()
        # Cleanup
        Path(expected_output).unlink()


# --- Test extended formats ---


class TestApplyParquet:
    def test_parquet_apply(self, temp_dir, cleaning_functions_file):
        pytest.importorskip("pyarrow")
        import pyarrow as pa
        import pyarrow.parquet as pq

        # Create test parquet file
        input_path = temp_dir / "data.parquet"
        table = pa.Table.from_pylist([
            {"name": "alice", "age": 30},
            {"name": "bob", "age": 25},
        ])
        pq.write_table(table, input_path)

        output = temp_dir / "output.parquet"
        result = apply_cleaning(str(input_path), str(cleaning_functions_file), str(output))

        assert result == str(output)
        result_table = pq.read_table(output)
        records = result_table.to_pylist()
        assert records[0]["name"] == "ALICE"

    def test_parquet_default_output(self, temp_dir, cleaning_functions_file):
        pytest.importorskip("pyarrow")
        import pyarrow as pa
        import pyarrow.parquet as pq

        input_path = temp_dir / "data.parquet"
        table = pa.Table.from_pylist([{"name": "test"}])
        pq.write_table(table, input_path)

        result = apply_cleaning(str(input_path), str(cleaning_functions_file))
        assert result == str(temp_dir / "data.cleaned.parquet")
        Path(result).unlink()


class TestApplyExcel:
    def test_xlsx_apply(self, temp_dir, cleaning_functions_file):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        # Create test xlsx file
        input_path = temp_dir / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["alice", 30])
        ws.append(["bob", 25])
        wb.save(input_path)

        output = temp_dir / "output.xlsx"
        result = apply_cleaning(str(input_path), str(cleaning_functions_file), str(output))

        assert result == str(output)

        from openpyxl import load_workbook
        wb_out = load_workbook(output)
        ws_out = wb_out.active
        assert ws_out.cell(2, 1).value == "ALICE"

    def test_xlsx_default_output(self, temp_dir, cleaning_functions_file):
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        input_path = temp_dir / "data.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name"])
        ws.append(["test"])
        wb.save(input_path)

        result = apply_cleaning(str(input_path), str(cleaning_functions_file))
        assert result == str(temp_dir / "data.cleaned.xlsx")
        Path(result).unlink()


class TestApplyText:
    def test_txt_to_md(self, temp_dir):
        # Create a text cleaning functions file
        code = '''
def clean_data(text):
    """Clean text by stripping whitespace."""
    return text.strip().upper()
'''
        func_file = temp_dir / "text_cleaner.py"
        func_file.write_text(code)

        input_path = temp_dir / "doc.txt"
        input_path.write_text("  hello world  ")

        output = temp_dir / "output.md"
        result = apply_cleaning(str(input_path), str(func_file), str(output))

        assert result == str(output)
        assert output.read_text() == "HELLO WORLD"

    def test_txt_default_output_is_md(self, temp_dir):
        code = 'def clean_data(text): return text'
        func_file = temp_dir / "cleaner.py"
        func_file.write_text(code)

        input_path = temp_dir / "doc.txt"
        input_path.write_text("test")

        result = apply_cleaning(str(input_path), str(func_file))
        assert result == str(temp_dir / "doc.cleaned.md")
        Path(result).unlink()


class TestGetDefaultOutputPathExtended:
    def test_xls_outputs_xlsx(self):
        result = get_default_output_path("data.xls")
        assert result == "data.cleaned.xlsx"

    def test_txt_outputs_md(self):
        result = get_default_output_path("doc.txt")
        assert result == "doc.cleaned.md"

    def test_pdf_outputs_md(self):
        result = get_default_output_path("doc.pdf")
        assert result == "doc.cleaned.md"

    def test_docx_outputs_md(self):
        result = get_default_output_path("doc.docx")
        assert result == "doc.cleaned.md"

    def test_parquet_stays_parquet(self):
        result = get_default_output_path("data.parquet")
        assert result == "data.cleaned.parquet"

    def test_xlsx_stays_xlsx(self):
        result = get_default_output_path("data.xlsx")
        assert result == "data.cleaned.xlsx"


# --- Test TUI colorization ---


class TestTUIColorization:
    def test_colorize_transmission_exists(self):
        """Verify colorize_transmission method exists."""
        pytest.importorskip("rich")
        from recursive_cleaner.tui import TUIRenderer

        renderer = TUIRenderer("test.jsonl", 10)
        assert hasattr(renderer, "_colorize_transmission")

    def test_colorize_with_issues(self):
        """Test colorization with issue elements."""
        pytest.importorskip("rich")
        from rich.text import Text
        from recursive_cleaner.tui import TUIRenderer

        renderer = TUIRenderer("test.jsonl", 10)
        response = '''
        <cleaning_analysis>
            <issues_detected>
                <issue id="1" solved="true">Phone format</issue>
                <issue id="2" solved="false">Date format</issue>
            </issues_detected>
            <chunk_status>needs_more_work</chunk_status>
        </cleaning_analysis>
        '''
        result = renderer._colorize_transmission(response)
        assert isinstance(result, Text)
        # Should contain text about issues
        plain = result.plain
        assert "Phone format" in plain or "ISSUES" in plain

    def test_colorize_with_function(self):
        """Test colorization with function generation."""
        pytest.importorskip("rich")
        from rich.text import Text
        from recursive_cleaner.tui import TUIRenderer

        renderer = TUIRenderer("test.jsonl", 10)
        response = '''
        <function_to_generate>
            <name>normalize_dates</name>
            <docstring>Convert dates to ISO format</docstring>
        </function_to_generate>
        <chunk_status>clean</chunk_status>
        '''
        result = renderer._colorize_transmission(response)
        assert isinstance(result, Text)
        plain = result.plain
        assert "normalize_dates" in plain or "GENERATING" in plain

    def test_colorize_status_clean(self):
        """Test clean status is handled."""
        pytest.importorskip("rich")
        from recursive_cleaner.tui import TUIRenderer

        renderer = TUIRenderer("test.jsonl", 10)
        response = '<chunk_status>clean</chunk_status>'
        result = renderer._colorize_transmission(response)
        assert "CLEAN" in result.plain.upper()

    def test_colorize_status_needs_more_work(self):
        """Test needs_more_work status is handled."""
        pytest.importorskip("rich")
        from recursive_cleaner.tui import TUIRenderer

        renderer = TUIRenderer("test.jsonl", 10)
        response = '<chunk_status>needs_more_work</chunk_status>'
        result = renderer._colorize_transmission(response)
        assert "NEEDS" in result.plain.upper() or "MORE" in result.plain.upper()
